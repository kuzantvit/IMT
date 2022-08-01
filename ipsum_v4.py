#/usr/share/python3/runtime.d/dh-python.rtupdate
#curl -X GET --header 'Accept: application/json' -u {6f9c962e-6e19-43dc-b8f1-788fccab72af:56ee7ec0-d1cf-4f8e-ab81-5034c7339223} 'https://exchange.xforce.ibmcloud.com/api/url/zmail.ru'

"curl -X GET --header 'Accept: application/json' -u {6f9c962e-6e19-43dc-b8f1-788fccab72af:56ee7ec0-d1cf-4f8e-ab81-5034c7339223} 'https://exchange.xforce.ibmcloud.com/api/ipr/89.234.157.254'"


import os
os.system("curl --compressed https://raw.githubusercontent.com/stamparm/ipsum/master/ipsum.txt > ipsum_table.txt")
import shodan
import csv
import getpass
import sys
import time
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Color, colors, PatternFill, Alignment
from openpyxl.styles.borders import Border, Side
import os
import json
import requests
import time
SHODAN_API_KEY = "N8s1tLahPUwQ5Z4ZO8ZJdeckR4YprHTn"
api = shodan.Shodan(SHODAN_API_KEY)

test =[]
arr_to_excel = []
ip = []
bad_count = []
score =[]
reason = []
shod_tags = []
alien_tags = []
vt_tags = []
raw_data = open(file='/usr/anton/raw.txt', mode ='a')
from OTXv2 import OTXv2
otx = OTXv2("9dca3e032fd9413fcc3ed9c680168d7f2ce3a546c79ea069fbec9a30a34dfb73")
import IndicatorTypes
with open('/usr/anton/ipsum_table.txt') as File:
#with open('U:\\scripts\\ipsum2.txt') as File:
    for line in File:
        test.append(line)
        #print(line)
for i in range(4, len(test)):
    a = test[i].split('\t')
    b = a[1].replace('\n', '')
    if int(b) > 10:
        ip.append(a[0])
        bad_count.append(b)
        print(a[0], type(a[0]))
        try:
            request = "curl -X GET --header \'Accept: application/json\' -u 6f9c962e-6e19-43dc-b8f1-788fccab72af:56ee7ec0-d1cf-4f8e-ab81-5034c7339223 \'https://exchange.xforce.ibmcloud.com/api/ipr/{}\' > ip_test1.txt".format(a[0])
            os.system(request)
            with open('/usr/anton/ip_test1.txt') as File:
                data = json.load(File)
                score.append(data['score'])
                aa= []
                for i in data['history']:
                    if '2019' in str(i['created']) or '2020' in str(i['created']):
                        aa.append('IP created: ' + str(i['created']) + ' and ' + str(i['reasonDescription']))
                reason.append('\n'.join(aa))
            os.remove('/usr/anton/ip_test1.txt')
        except:
            score.append('No Info')
            reason.append('No Info')
        try:
            host_shodan = api.host(a[0])
            shod_tags.append(';\t'.join(host_shodan['tags']))
        except:
            shod_tags.append('No Info')
        try:
            result = otx.get_indicator_details_by_section(IndicatorTypes.IPv4, a[0], 'general')
            otx_test= []
            for i in result['pulse_info']['pulses']:
                otx_test.append(str(i['name']) + ' ' + str(i['tags']))
            alien_tags.append('\n'.join(otx_test))
        except Exception as e:
            print(e)
            alien_tags.append('No Info')
        try:
            simple_vt = []
            url = 'https://www.virustotal.com/vtapi/v2/ip-address/report'
            params = {'apikey':'6ff17d7b37810ba16a3d817e4dcc91e211483d33d00bbffc15240c5938f66c2b','ip':str(a[0])}
            response = requests.get(url, params=params)
            data_vt = response.json()
            data_v_t = data_vt['detected_urls']
            for i in data_v_t:
                if '2019' in str(i['scan_date']) or '2020' in str(i['scan_date']):
                    simple_vt.append(' Bad result: ' + str(i['positives']) + ' Date:' + str(i['scan_date'])+ ' Url: ' + str(i['url']))
            vt_tags.append('\n'.join(simple_vt))
        except Exception as e:
            print(e)
            vt_tags.append('No Info')
        raw_data.write(a[0] + ';' + b + ';' + str(score[-1]) + ';'+ str(reason[-1]) + ';' + str(shod_tags[-1]) + ';' + str(alien_tags[-1]) + ';'+ str(vt_tags[-1]) +'\n')
        time.sleep(14)

raw_data.close()
def save_data_to_excel(*args, filename='/usr/anton/ipsum.xlsx', sheet_name='Main', **kwargs):
    fill_cell_orange = PatternFill(start_color='f07233',
               end_color='f07233',
               fill_type='solid')
    alignment_cell = Alignment(horizontal='left', vertical ='center', wrapText='False')

    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
    try:
        wb_write = load_workbook(filename)
    except FileNotFoundError:
        wb_write = Workbook()
    ws_write = wb_write.create_sheet(sheet_name)
    for counter, k in enumerate(kwargs):
        #print(counter)
        #print(k)
        #print(type(k))
        ws_write.cell(row=1, column=counter+1).value = kwargs[k]
    for count_data, data in enumerate(args):
        for counter, info in enumerate(data):
            ws_write.cell(row=counter+2, column=count_data+1).value = data[counter]
            ws_write.cell(row=counter+2, column=count_data+1).alignment = alignment_cell
            ws_write.cell(row=counter+2, column=count_data+1).border = thin_border
            try:
                if 'poweredOff' in (data[counter]):
                    ws_write.cell(row=counter+2, column=count_data+1).fill = fill_cell_orange
            except:
                pass
    wb_write.save(filename)
print(ip, bad_count, score)
save_data_to_excel(ip, bad_count, score, reason, shod_tags, alien_tags, vt_tags, filename='/usr/anton/ipsum.xlsx')
print('finish')

