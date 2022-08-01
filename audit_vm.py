from pyVim import connect
from pyVim.connect import Disconnect
import ssl
from pyVmomi import vim
import csv
import getpass
import sys
import time
import re
import getpass
import time
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Color, colors, PatternFill, Alignment
from openpyxl.styles.borders import Border, Side
import os
import configparser

# Исключения по именам ВМ и папок в которых они находятся добавлены в файл конфигурации, новые данные вносить через запятую без пробела
config = configparser.ConfigParser()
config.read(r'G:\Департамент информационной безопасности\14_ОЭСЗИ\1_Отчеты\АУДИТ ВМ\config\config.ini')
try:
    os.remove("U:\\отчеты\\audit_vm_test\\audit_vm.xlsx")
except:
    pass
try:
    os.remove("U:\\отчеты\\audit_vm_test\\audit_vmware.xlsx")
except:
    pass

ssl_context = ssl.SSLContext(ssl.PROTOCOL_SSLv23)
ssl_context.verify_mode = ssl.CERT_NONE
hosts = ['vc-res-01.sibintek.ru', 'vc-vdi-p1-01.sibintek.ru', 'vc-vdi-p2-01.sibintek.ru', 'vc-mgmt-01.sibintek.ru']
#hosts = ['vc-res-01.sibintek.ru', 'vc-mgmt-01.sibintek.ru']
#hosts = ['vc-mgmt-01.sibintek.ru']

# Ввод логина и пароля для подключения к VMware

username = input('Username:')
password = getpass.getpass()
print(password)
assert len(password) > 0

try :
    if len(password) < 5 or password == username:
        raise Exception
    else:
        pass
except Exception as e:
    print("Пароль не удалось скопировать")
    time.sleep(6)
'''
vm_orig_name = []
vm_name = []
vm_power = []
vm_ip = []
vm_type = []
vm_saas = []
vm_meta_name = []
ksv_status = []

vm_agent = []
vm_soft_name = []
vm_sec_name = []

ksv_status_agent =[]
ksv_status_soft_name = []
ksv_status_agent_sec = []
'''
counter=0


#Функция для подключения в варе и сбора данных 

def get_data_vsphere(hosts, password, username):
    vm_orig_name = []
    vm_name = []
    vm_power = []
    vm_ip = []
    vm_annotation = []
    vm_type = []
    vm_saas = []
    vm_meta_name = []
    vm_orig_name_sheet_2 = []
    vm_name_sheet_2 = []
    vm_power_sheet_2 = []
    vm_ip_sheet_2 = []
    vm_annotation_sheet_2 = []
    vm_type_sheet_2 = []
    vm_saas_sheet_2 = []
    vm_meta_name_sheet_2 = []
    vm_host_name = []
    vm_host_name_sheet_2 = []
    vmm_os_name = []
    vmm_os_name_sheet_2 = []
    '''except_vms = ['NSX_Controller_2762c66b-10e0-4349-94d4-cf3235b09579', 'NSX_Controller_5f8161c4-5fbe-4fb0-a436-b90704fe01e4', 'NSX_Controller_da7667b8-3c9e-48f4-a32e-d4bc236b880d',\
                  'nsx-mgr-mgmt-01', 'psc-mgmt-01.sibintek.ru', 'vc-mgmt-01.sibintek.ru', 'vrl-cm-01.sibintek.ru', 'usage-meter.sibintek.ru', 'idm-mgr-01.sibintek.ru',\
                  'idm-mgr-02.sibintek.ru', 'idm-mgr-03.sibintek.ru', 'nsx-mgr-vdi-p2-01', 'nsx-mgr-vdi-p1-01', 'vc-vdi-p2-01.sibintek.ru', 'vc-vdi-p1-01.sibintek.ru',\
                  'vdi-p2-uag1-01.sibintek.ru', 'vdi-p2-uag2-01.sibintek.ru', 'vdi-p2-uag3-01.sibintek.ru', 'vdi-p1-uag1-01.sibintek.ru', 'vdi-p1-uag2-01.sibintek.ru',\
                  'vdi-p1-uag3-01.sibintek.ru', 'psc-res-02.sibintek.ru', 'nsx-mgr-res-01', 'psc-res-01.sibintek.ru', 'vc-res-01.sibintek.ru', 'vs-sib-mgt-eps',\
                  'rds-p1-uag3-01.sibintek.ru', 'rds-p1-uag2-01.sibintek.ru', 'rds-p1-uag1-01.sibintek.ru', 'CLONE-cdc-rdgw-001', 'sib-netapp-dbrk.sibintek.ru', 'vdi-p2-uag1-01.sibintek.ru (new)',\
                  'sensor.sibintek.ru', 'sensor-prevent.sibintek.ru', 'decryptor.sibintek.ru', 'huntbox.sibintek.ru']
    '''
    #Создание массивов с исключениями
    except_vms = config.get('Exceptions', 'vmname').split(',')
    except_parent_name = config.get('Exceptions', 'parentname').split(',')
    for host_name in hosts:
        print('Getting data from: ' + str(host_name))
        #print(password)
        try:
            si = connect.SmartConnect(host=host_name, user=username, pwd=password, port=443, sslContext=ssl_context) # подключение к соответствующему кластеру
            content = si.RetrieveContent()
            container = content.rootFolder
            viewType = [vim.VirtualMachine]
            recursive = True
            containerView = content.viewManager.CreateContainerView(container, viewType, recursive)

            children = containerView.view
        except Exception as e:
            print(e)
            pass
# цикл по выгрузке информации по ВМ, сразу идет удаление домена, создание массива по сервисной модели, исключение делается по ппаке в которой лежит ВМ, её имени, кластеру
        for vm in children: 
            if vm.name is not None and vm.name not in except_vms and vm.parent.name not in except_parent_name and host_name!='vc-vdi-p1-01.sibintek.ru' and host_name != 'vc-vdi-p2-01.sibintek.ru':
                vm_name.append(vm.summary.guest.hostName.lower().replace('.sibintek.ru', '').replace('.cosn.cdc', '').replace('.rnbvk.ru', '') if vm.summary.guest.hostName is not None else vm.summary.guest.hostName)
                vm_power.append(vm.summary.runtime.powerState)
                vm_orig_name.append(vm.name.lower().replace('.sibintek.ru', '').replace('.cosn.cdc', '').replace('.rnbvk.ru', '') if vm.name is not None else vm.name)
                vm_host_name.append(host_name)
                vm_ip.append(vm.summary.guest.ipAddress)
                vm_annotation.append(vm.config.annotation)
                result_meta = (str(vm.summary.vm).replace("vim.VirtualMachine:", '')).replace("'", '')
                vm_meta_name.append(result_meta)
                vmm_os_name.append(vm.guest.guestFullName)
                #Распределение данных ВМ по группам СааС и ИааС в зависимости от имени,кластера и папке в которой они лежат
                if vm.parent.name not in ['SaaS', 'IaaS']:
                    if vm.parent.parent.name == 'KIS-1C-IAAS (150e34b5-78be-4ff8-b40a-84d99756383f)' or 'KIS-1C-IAAS' in vm.parent.parent.name:
                        vm_saas.append('IaaS')
                        vm_type.append(vm.parent.name)
                        #print('test1')
                    elif vm.parent.parent.name == 'RDSE (VIEWPLANNER)':
                        vm_saas.append('SaaS')
                        vm_type.append(vm.parent.name)
                        #print('test2')
                    elif vm.parent.parent.name == 'Security':
                        vm_saas.append('SaaS')
                        vm_type.append(vm.parent.name)
                        #print('test3')
                    elif vm.parent.parent.name == 'ViewPlanner':
                        vm_saas.append('SaaS')
                        vm_type.append(vm.parent.name)
                        #print('test4')
                    elif vm.parent.parent.name == 'RDS':
                        vm_saas.append('SaaS')
                        vm_type.append(vm.parent.name)
                        #print('test5')
                    elif host_name == 'vc-mgmt-01.sibintek.ru':
                        vm_saas.append('SaaS')
                        vm_type.append(vm.parent.name)
                        #print('test6')
                    elif vm.name == 'RaaSExch2012r2':
                        vm_saas.append('SaaS')
                        vm_type.append(vm.parent.name)
                        #print('test7')
                    elif vm.name == 'CDC-DNS-02' or vm.name == 'CDC-DNS-01':
                        vm_saas.append('SaaS')
                        vm_type.append(vm.parent.name)
                        #print('test8')
                    elif vm.parent.parent.name == 'horizon':
                        vm_saas.append('SaaS')
                        vm_type.append(vm.parent.name)
                    else:
                        vm_saas.append(vm.parent.parent.name)
                        vm_type.append(vm.parent.name)
                        #print('test9')
                else:
                    if host_name == 'vc-mgmt-01.sibintek.ru':
                        vm_saas.append('SaaS')
                        vm_type.append('No specification')
                        #print('test10')
                    else:
                        vm_saas.append(vm.parent.name)
                        vm_type.append('No specification')
            # Размещение ВМ во отдельные массивы которые пойдут потом во вклдадку Extra
            elif vm.name is not None and (vm.name in except_vms or vm.parent.name in except_parent_name or host_name=='vc-vdi-p1-01.sibintek.ru' or host_name == 'vc-vdi-p2-01.sibintek.ru' or 'svm-la-hv' in vm.name):
                vm_name_sheet_2.append(vm.summary.guest.hostName.lower() if vm.summary.guest.hostName is not None else vm.summary.guest.hostName)
                vm_power_sheet_2.append(vm.summary.runtime.powerState)
                vm_orig_name_sheet_2.append(vm.name.lower() if vm.name is not None else vm.name)
                vm_host_name_sheet_2.append(host_name)
                vm_ip_sheet_2.append(vm.summary.guest.ipAddress)
                vm_annotation_sheet_2.append(vm.config.annotation)
                result_meta_2 = (str(vm.summary.vm).replace("vim.VirtualMachine:", '')).replace("'", '')
                vm_meta_name_sheet_2.append(result_meta_2)
                vmm_os_name_sheet_2.append(vm.guest.guestFullName)
                if vm.parent.name not in ['SaaS', 'IaaS']:
                    vm_saas_sheet_2.append(vm.parent.parent.name)
                    vm_type_sheet_2.append(vm.parent.name)
                else:
                    vm_saas_sheet_2.append(vm.parent.name)
                    vm_type_sheet_2.append('No specification')
    Disconnect(si)
    status_vm = ['vmware'] * len(vm_name)
    
    status_vm_sheet_2 = ['vmware'] * len(vm_name_sheet_2)
    
    return vm_name, vm_power, vm_orig_name, vm_ip, vm_meta_name, vm_annotation, vm_saas, vm_type, status_vm, vm_host_name, vmm_os_name, vm_name_sheet_2,\
           vm_power_sheet_2, vm_orig_name_sheet_2, vm_ip_sheet_2, vm_meta_name_sheet_2, vm_annotation_sheet_2, vm_saas_sheet_2, vm_type_sheet_2, status_vm_sheet_2, vm_host_name_sheet_2, vmm_os_name_sheet_2
# вызов выше описанной функции
(vm_name, vm_power, vm_orig_name, vm_ip, vm_meta_name, vm_annotation, vm_saas, vm_type, status_vm, vm_host_name, vmm_os_name, vm_name_sheet_2,\
 vm_power_sheet_2, vm_orig_name_sheet_2, vm_ip_sheet_2, vm_meta_name_sheet_2, vm_annotation_sheet_2, vm_saas_sheet_2,\
 vm_type_sheet_2, status_vm_sheet_2, vm_host_name_sheet_2, vmm_os_name_sheet_2) = get_data_vsphere(hosts, password, username)
#Функция для сравнения выгруженных ВМ с отчетами касперского в самом отчете отрубаются домены для сравнения с данными из касперского    
def check_ksv_files(vm_meta_name, vm_orig_name, vm_name):
    if vm_meta_name == [im for im in range(len(vm_orig_name))]:
        ksv_dict ={}
        ksv_status = []
        for i in range(len(vm_name)):
            ksv_dict.update({vm_meta_name[i]:'NotProtected'}) 
        for v_1 in vm_meta_name:
            #print(ksv_dict.get(vm_meta_name))
            ksv_status.append('NotProtected')

    else:
        ksv_dict ={}        
        ksv_status = []    
        with open('U:\\scripts\\vmware.csv', newline='') as File:  
            reader = csv.reader(File)
            for row in reader:
                ksv_dict.update({row[0]:row[2]})
        with open('U:\\scripts\\vmware_mgmt.csv', newline='') as File:  
            reader = csv.reader(File)
            for row in reader:
                ksv_dict.update({row[0]:row[2]})
        for v_1 in vm_meta_name:
            #print(ksv_dict.get(vm_meta_name))
            ksv_status.append(ksv_dict.get(v_1))

    #print(ksv_dict)
    ksv_status_agent =[]
    ksv_status_soft_name = []
    ksv_status_agent_sec = []
    wb_ksv = load_workbook('U:\\scripts\\ksc_hyperv.xlsx')
    ws_ksv = wb_ksv['Details']
    vm_agent_dict ={}
    vm_soft_name_dict = {}
    vm_sec_name_dict = {}
    for i in range(len(vm_orig_name)):
        for m in range(2666):
            if vm_orig_name[i] is not None and ws_ksv.cell(row=m+1,column=3).value is not None and vm_name[i] is not None and (ksv_dict.get(vm_meta_name[i]) == 'NotProtected' or ksv_dict.get(vm_meta_name[i]) is None) and\
               (vm_name[i].lower().replace('.sibintek.ru', '').replace('.rnbvk.ru', '').replace('.cosn.cdc', '') ==\
                ws_ksv.cell(row=m+1,column=3).value.lower().replace('.sibintek.ru', '').replace('.rnbvk.ru', '').replace('.cosn.cdc', '') or\
                vm_orig_name[i].lower().replace('.sibintek.ru', '').replace('.cosn.cdc', '').replace('.rnbvk.ru', '') ==\
                ws_ksv.cell(row=m+1,column=3).value.lower().replace('.sibintek.ru', '').replace('.rnbvk.ru', '').replace('.cosn.cdc', '')):
                vm_agent_dict.update({vm_orig_name[i]:ws_ksv.cell(row=m+1,column=4).value})
                vm_soft_name_dict.update({vm_orig_name[i]:ws_ksv.cell(row=m+1,column=5).value})
                vm_sec_name_dict.update({vm_orig_name[i]:ws_ksv.cell(row=m+1,column=6).value})
            elif vm_name[i] is None and vm_orig_name[i] is not None and ws_ksv.cell(row=m+1,column=3).value is not None and (ksv_dict.get(vm_meta_name[i]) == 'NotProtected' or ksv_dict.get(vm_meta_name[i]) is None) and\
                 vm_orig_name[i].lower().replace('.sibintek.ru', '').replace('.cosn.cdc', '').replace('.rnbvk.ru', '') ==\
                 ws_ksv.cell(row=m+1,column=3).value.lower().replace('.sibintek.ru', '').replace('.rnbvk.ru', '').replace('.cosn.cdc', ''):
                vm_agent_dict.update({vm_orig_name[i]:ws_ksv.cell(row=m+1,column=4).value})
                vm_soft_name_dict.update({vm_orig_name[i]:ws_ksv.cell(row=m+1,column=5).value})
                vm_sec_name_dict.update({vm_orig_name[i]:ws_ksv.cell(row=m+1,column=6).value})
                
    wb_ksv.save('U:\\scripts\\ksc_hyperv.xlsx')

    wb_ksv = load_workbook('U:\\scripts\\ksc_vmware.xlsx')
    ws_ksv = wb_ksv['Details']
    for i in range(len(vm_orig_name)):
        for m in range(2066):
            if vm_orig_name[i] is not None and ws_ksv.cell(row=m+1,column=3).value is not None and vm_name[i] is not None and (ksv_dict.get(vm_meta_name[i]) == 'NotProtected' or ksv_dict.get(vm_meta_name[i]) is None) and\
               (vm_name[i].lower().replace('.sibintek.ru', '').replace('.rnbvk.ru', '').replace('.cosn.cdc', '') ==\
                ws_ksv.cell(row=m+1,column=3).value.lower().replace('.sibintek.ru', '').replace('.rnbvk.ru', '').replace('.cosn.cdc', '') or\
                vm_orig_name[i].lower().replace('.sibintek.ru', '').replace('.cosn.cdc', '').replace('.rnbvk.ru', '') ==\
                ws_ksv.cell(row=m+1,column=3).value.lower().replace('.sibintek.ru', '').replace('.rnbvk.ru', '').replace('.cosn.cdc', '')):
                #print(vm_orig_name[i], vm_meta_name[i])
                vm_agent_dict.update({vm_orig_name[i]:ws_ksv.cell(row=m+1,column=4).value})
                vm_soft_name_dict.update({vm_orig_name[i]:ws_ksv.cell(row=m+1,column=5).value})
                vm_sec_name_dict.update({vm_orig_name[i]:ws_ksv.cell(row=m+1,column=6).value})
            elif vm_name[i] is None and vm_orig_name[i] is not None and ws_ksv.cell(row=m+1,column=3).value is not None and (ksv_dict.get(vm_meta_name[i]) == 'NotProtected' or ksv_dict.get(vm_meta_name[i]) is None) and\
                 vm_orig_name[i].lower().replace('.sibintek.ru', '').replace('.cosn.cdc', '').replace('.rnbvk.ru', '') ==\
                 ws_ksv.cell(row=m+1,column=3).value.lower().replace('.sibintek.ru', '').replace('.rnbvk.ru', '').replace('.cosn.cdc', ''):
                vm_agent_dict.update({vm_orig_name[i]:ws_ksv.cell(row=m+1,column=4).value})
                vm_soft_name_dict.update({vm_orig_name[i]:ws_ksv.cell(row=m+1,column=5).value})
                vm_sec_name_dict.update({vm_orig_name[i]:ws_ksv.cell(row=m+1,column=6).value})    
    for v_2 in vm_orig_name:
        #print(ksv_dict.get(vm_meta_name))
        ksv_status_agent.append(vm_agent_dict.get(v_2))
        ksv_status_soft_name.append(vm_soft_name_dict.get(v_2))
        ksv_status_agent_sec.append(vm_sec_name_dict.get(v_2))
    wb_ksv.save('U:\\scripts\\ksc_vmware.xlsx')
        
    return ksv_status, ksv_status_agent, ksv_status_soft_name, ksv_status_agent_sec

ksv_status, ksv_status_agent, ksv_status_soft_name, ksv_status_agent_sec=check_ksv_files(vm_meta_name, vm_orig_name, vm_name)
ksv_status_sheet_2, ksv_status_agent_sheet_2, ksv_status_soft_name_sheet_2, ksv_status_agent_sec_sheet_2 =check_ksv_files(vm_meta_name_sheet_2, vm_orig_name_sheet_2, vm_name_sheet_2,)

# Проверка получения данных т.к. все лежит в разных массивах (не удачное решение) то размер массивов должен совпадать, чтобы данные не переехали
print(len(vm_name), len(vm_orig_name), len(vm_power), len(vm_ip), len(vm_meta_name), len(vm_saas), len(vm_type), len(ksv_status), len(ksv_status_agent), len(ksv_status_soft_name), len(ksv_status_agent_sec))
print(type(vm_name), type(vm_power), type(vm_ip), type(vm_meta_name), type(vm_saas), type(vm_type), type(ksv_status), type(ksv_status_agent), type(ksv_status_soft_name), type(ksv_status_agent_sec))


# Определение текущей даты, не уверене, что в скрипте это используется
import datetime


now = datetime.datetime.now()
if now.month < 10:
    month_string = '0' + str(now.month)
else:
    month_string = str(now.month)

if now.day < 10:
    day_string = '0' + str(now.day)
else:
    day_string = str(now.day)
date_string = day_string + month_string
fill_cell_orange = PatternFill(start_color='f07233',
               end_color='f07233',
               fill_type='solid')
# Это создание атрибутов для excel(цвета ячеек)
fill_cell_green = PatternFill(start_color='05f621', end_color='05f621', fill_type='solid')
fill_cell_red = PatternFill(start_color='f60511', end_color='f60511', fill_type='solid')
fill_cell_yellow = PatternFill(start_color='ecee17', end_color='ecee17', fill_type='solid')
fill_cell_red_dns = PatternFill(start_color='fa0c79', end_color='f60511', fill_type='solid')
fill_cell_green_dns = PatternFill(start_color='52df5b', end_color='05f621', fill_type='solid')
fill_cell_yellow_dns = PatternFill(start_color='f6f875', end_color='ecee17', fill_type='solid')


'''
for child in children:
    print(child.config.)

'''
# функция для сохранения данных в excel
def save_data_to_excel(*args, filename='U:\\отчеты\\audit_vm_test\\audit.xlsx', sheet_name='Main', **kwargs):
    fill_cell_orange = PatternFill(start_color='f07233',
               end_color='f07233',
               fill_type='solid')
    alignment_cell = Alignment(horizontal='left', vertical ='center', wrapText='False')

    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
    try:
        wb_write = load_workbook(filename)
    except FileNotFoundError:
        wb_write = Workbook()
    ws_write = wb_write.create_sheet(sheet_name)
    for counter, k in enumerate(kwargs):
        #print(counter)
        #print(k)
        #print(type(k))
        ws_write.cell(row=1, column=counter+1).value = kwargs[k]
    for count_data, data in enumerate(args):
        for counter, info in enumerate(data):
            ws_write.cell(row=counter+2, column=count_data+1).value = data[counter]
            ws_write.cell(row=counter+2, column=count_data+1).alignment = alignment_cell
            ws_write.cell(row=counter+2, column=count_data+1).border = thin_border
            try:
                if 'poweredOff' in (data[counter]):
                    ws_write.cell(row=counter+2, column=count_data+1).fill = fill_cell_orange
            except:
                pass
    wb_write.save(filename)
    
save_data_to_excel(vm_orig_name, vm_name, vm_meta_name, ksv_status, ksv_status_agent, ksv_status_soft_name, ksv_status_agent_sec, vm_power, vm_ip, vm_saas,\
                   vm_type, status_vm, vmm_os_name, vm_host_name, vm_annotation, filename='U:\\отчеты\\audit_vm_test\\audit_vmware.xlsx',sheet_name='Main', vm_or = "VM Name", vm_dns='DNS Name',\
                   vmmetaname='VM Meta Name', ksv_status_n='Kaspersky Status', ksv_status_agentn='Версия Агента администрирования',\
                   ksv_status_softn='Название программы безопасности', ksv_status_dbsoftn='Версия программы безопасности',\
                   runstate = 'Power State', ip_add = 'IP address')

save_data_to_excel(vm_orig_name_sheet_2, vm_name_sheet_2, vm_meta_name_sheet_2, ksv_status_sheet_2, ksv_status_agent_sheet_2,\
                   ksv_status_soft_name_sheet_2, ksv_status_agent_sec_sheet_2, vm_power_sheet_2, vm_ip_sheet_2, vm_saas_sheet_2,\
                   vm_type_sheet_2, status_vm_sheet_2, vmm_os_name_sheet_2, vm_host_name_sheet_2, vm_annotation_sheet_2,\
                   filename='U:\\отчеты\\audit_vm_test\\audit_vmware.xlsx', sheet_name='extra',\
                   vm_or = "VM Name", vm_dns='DNS Name',\
                   vmmetaname='VM Meta Name', ksv_status_n='Kaspersky Status', ksv_status_agentn='Версия Агента администрирования',\
                   ksv_status_softn='Название программы безопасности', ksv_status_dbsoftn='Версия программы безопасности',\
                   runstate = 'Power State', ip_add = 'IP address')

#save_data_to_excel(vm_orig_name, vm_name, vm_meta_name, ksv_status, ksv_status_agent, ksv_status_soft_name, ksv_status_agent_sec, vm_power, vm_ip, vm_saas, vm_type, filename='U:\\отчеты\\audit_vm_test\\audit_vm_vmware.xlsx')

'''
def get_data_from_vmmreport(filename='U:\\scripts\\vmmreport.csv'):
    vmm_name = []
    vmm_cr_date ={}
    vmm_ch_date ={}
    vmm_status = {}
    vmm_cloud = {}
    vmm_service = {}
    vmm_dns_name = []
    vmm_os_name = {}
    counter = 0
    with open('U:\\scripts\\vmmreport.csv', newline='') as File:  
        reader = csv.reader(File, delimiter=';')
        for row in reader:
            if 'Name' in row:
                for i in range(len(row)):
                    if row[i] == 'Name':
                        Name = i
                    if row[i] == 'Status':
                        Status = i
                    if row[i] == 'AddedTime':
                        cr_date = i
                    if row[i] == 'ModifiedTime':
                        ch_date = i
                    if row[i] == 'Cloud':
                        cloud = i
                    
                    if row[i] == 'ComputerName':
                        dns_name = i
                    if row[i] == 'OperatingSystem':
                        os_name = i
                    
                continue
            counter+=1
            #print(row)
            vmm_name.append(row[Name])
            vmm_dns_name.append(row[dns_name])
            if row[cloud] in ['Sibintek_ServiceVM', 'Sibintek_RDS',\
                          'Sibintek_RNBVK', 'COSN_Management', 'Sibintek_NOC',\
                          'Sibintek_AD', 'Sibintek_DIB', 'COSN_NIX', 'Sibintek_NIX', 'Sibintek_SaaS']:
                vmm_service.update({row[Name]:'SaaS'})
            elif row[cloud] in ['Sibintek_IaaS', 'Sibintek_IasS_NIX', 'Sibintek_DIRPP',\
                             'Sibintek_Zvezda', 'RN-Bitum', 'Sibintek_Invest', 'Sibintek_ITSM', 'Sibintek_1C']:
                vmm_service.update({row[Name]:'IaaS'})
            else:
                vmm_service.update({row[Name]:'Delete_VM'})
            vmm_cr_date.update({row[Name]:row[cr_date]})
            vmm_ch_date.update({row[Name]:row[ch_date]})
            status = ("poweredOn" if row[Status] =="Running" else 'poweredOff')
            vmm_status.update({row[Name]:status})
            vmm_cloud.update({row[Name]:row[cloud]})
            vmm_os_name.update({row[Name]:row[os_name]})
    return vmm_name, vmm_cr_date, vmm_ch_date, vmm_status, vmm_cloud, vmm_service, vmm_dns_name, vmm_os_name
'''
#Функция для чтения данных из отчета по cosn
def get_data_from_vmmreport(filename='U:\\scripts\\vmmreport.csv'):
    vmm_name = []
    vmm_cr_date =[]
    vmm_ch_date =[]
    vmm_status = []
    vmm_cloud = []
    vmm_service = []
    vmm_dns_name = []
    vmm_os_name = []
    counter = 0
    with open('U:\\scripts\\vmmreport.csv', newline='') as File:  
        reader = csv.reader(File, delimiter=';')
        for row in reader:
            if 'Name' in row:
                for i in range(len(row)):
                    if row[i] == 'Name':
                        Name = i
                    if row[i] == 'Status':
                        Status = i
                    if row[i] == 'AddedTime':
                        cr_date = i
                    if row[i] == 'ModifiedTime':
                        ch_date = i
                    if row[i] == 'Cloud':
                        cloud = i
                    
                    if row[i] == 'ComputerName':
                        dns_name = i
                    if row[i] == 'OperatingSystem':
                        os_name = i
                    
                continue
            counter+=1
            #print(row)
            vmm_name.append(row[Name])
            vmm_dns_name.append(row[dns_name])
            # Разбиение на СааС и ИааС тачек в косн
            if row[cloud] in ['Sibintek_ServiceVM', 'Sibintek_RDS',\
                          'COSN_Management', 'Sibintek_NOC',\
                          'Sibintek_AD', 'Sibintek_DIB', 'COSN_NIX', 'Sibintek_NIX', 'Sibintek_SaaS']:
                vmm_service.append('SaaS')
            elif row[cloud] in ['Sibintek_IaaS', 'Sibintek_IasS_NIX', 'Sibintek_DIRPP',\
                             'Sibintek_Zvezda', 'RN-Bitum', 'Sibintek_Invest', 'Sibintek_ITSM', 'Sibintek_1C', 'Sibintek_RNBVK']:
                vmm_service.append('IaaS')
            else:
                vmm_service.append('Delete_VM')
            vmm_cr_date.append(row[cr_date])
            vmm_ch_date.append(row[ch_date])
            status = ("poweredOn" if row[Status] =="Running" else 'poweredOff')
            vmm_status.append(status)
            vmm_cloud.append(row[cloud])
            vmm_os_name.append(row[os_name])
    return vmm_name, vmm_cr_date, vmm_ch_date, vmm_status, vmm_cloud, vmm_service, vmm_dns_name, vmm_os_name


vmm_name, vmm_cr_date, vmm_ch_date, vmm_status, vmm_cloud, vmm_service, vmm_dns_name, vmm_os_name = get_data_from_vmmreport()


#ksv_status, ksv_status_agent, ksv_status_soft_name, ksv_status_agent_sec=check_ksv_files(['test'], vmm_name)
ksv_status, ksv_status_agent, ksv_status_soft_name, ksv_status_agent_sec=check_ksv_files([im for im in range(len(vmm_name))], vmm_name, vmm_dns_name)
print('Hyper-V test')
print(len(ksv_status), len(ksv_status_agent), len(ksv_status_soft_name), len(ksv_status_agent_sec), len(vmm_dns_name))

# Внесение данных в отчет (итоговый) по косн
wb_ksv = load_workbook('U:\\отчеты\\audit_vm_test\\audit_vmware.xlsx')
ws_ksv = wb_ksv['Main']
ws_ksv_extra = wb_ksv['extra']
for i in range(len(vmm_name)):
    #print(vmm_name[i])
        #print(ws_ksv.cell(row=m+1, column=1).value)
        #print(vmm_name[i].lower())
    if 'ce-0009' not in vmm_name[i].lower() and 'la-svm-hv' not in vmm_name[i].lower() and vmm_service[i] != 'Delete_VM' and 'svm-la-hv' not in vmm_name[i].lower():
        for m in range(1700):
            if ws_ksv.cell(row=m+1, column=1).value is None:
                ws_ksv.cell(row=m+1, column=1).value = vmm_name[i].lower().replace('.sibintek.ru', '').replace('.cosn.cdc', '').replace('.rnbvk.ru', '')
                ws_ksv.cell(row=m+1, column=2).value = vmm_dns_name[i].lower().replace('.sibintek.ru', '').replace('.cosn.cdc', '').replace('.rnbvk.ru', '') if vmm_dns_name[i] is not None else vmm_dns_name[i]
                ws_ksv.cell(row=m+1, column=10).value = vmm_service[i]
                ws_ksv.cell(row=m+1, column=11).value = vmm_cloud[i]
                ws_ksv.cell(row=m+1, column=8).value = vmm_status[i]
                ws_ksv.cell(row=m+1, column=12).value = 'Hyper-V'
                ws_ksv.cell(row=m+1, column=5).value = ksv_status_agent[i]
                ws_ksv.cell(row=m+1, column=6).value = ksv_status_soft_name[i]
                ws_ksv.cell(row=m+1, column=7).value = ksv_status_agent_sec[i]
                ws_ksv.cell(row=m+1, column=13).value = vmm_os_name[i]
                ws_ksv.cell(row=m+1, column=16).value = vmm_cr_date[i]
                ws_ksv.cell(row=m+1, column=17).value = vmm_ch_date[i]
            else:
                continue
            break
    elif 'ce-0009' in vmm_name[i].lower() or 'la-svm-hv' in vmm_name[i].lower() or vmm_service[i] == 'Delete_VM' or 'svm-la-hv' in vmm_name[i].lower():
        print("Совпадение")
        for m in range(1700):
            if ws_ksv_extra.cell(row=m+1, column=1).value is None:   
                ws_ksv_extra.cell(row=m+1, column=1).value = vmm_name[i].lower()
                ws_ksv_extra.cell(row=m+1, column=2).value = vmm_dns_name[i].lower() if vmm_dns_name[i] is not None else vmm_dns_name[i]
                ws_ksv_extra.cell(row=m+1, column=10).value = vmm_service[i]
                ws_ksv_extra.cell(row=m+1, column=11).value = vmm_cloud[i]
                ws_ksv_extra.cell(row=m+1, column=8).value = vmm_status[i]
                ws_ksv_extra.cell(row=m+1, column=12).value = 'Hyper-V'
                ws_ksv_extra.cell(row=m+1, column=5).value = ksv_status_agent[i]
                ws_ksv_extra.cell(row=m+1, column=6).value = ksv_status_soft_name[i]
                ws_ksv_extra.cell(row=m+1, column=7).value = ksv_status_agent_sec[i]
                ws_ksv_extra.cell(row=m+1, column=13).value = vmm_os_name[i]
                ws_ksv_extra.cell(row=m+1, column=16).value = vmm_cr_date[i]
                ws_ksv_extra.cell(row=m+1, column=17).value = vmm_ch_date[i]
            else:
                continue
            break
            
            
wb_ksv.save('U:\\отчеты\\audit_vm_test\\audit_vmware.xlsx')

wb_ksv = load_workbook('U:\\отчеты\\audit_vm_test\\audit_vmware.xlsx')
ws_ksv = wb_ksv['Main']
for i in range(2000):
    if ws_ksv.cell(row=i+2,column=5).value != None and (ws_ksv.cell(row=i+2,column=6).value != 'N/A' and ws_ksv.cell(row=i+2,column=6).value != None):
        ws_ksv.cell(row=i+2,column=4).value = 'Protected'
    if ws_ksv.cell(row=i+2,column=1).value is not None and 'aspmx' in ws_ksv.cell(row=i+2,column=1).value:
        ws_ksv.cell(row=i+2,column=4).value = 'Protected'

# Проверка на дупликаты
vm_orig_name_dup = []
vm_orig_name_dup_exception = []
for i in range(2000):
    if ws_ksv.cell(row=i+2,column=1).value != None:
        vm_orig_name_dup.append(ws_ksv.cell(row=i+2,column=1).value)
#print(len(vm_orig_name_dup))


for vm in vm_orig_name_dup:
    if vm_orig_name_dup.count(vm) > 1 and vm not in vm_orig_name_dup_exception:
        print(vm)
        counter = 0
        for k in range(len(vm_orig_name_dup)):
            if ws_ksv.cell(row=k+2,column=1).value == vm and counter == 0:
                ws_ksv.cell(row=k+2,column=1).fill = fill_cell_red
                a_1,b_1,c_1,d_1 = ws_ksv.cell(row=k+2,column=4).value, ws_ksv.cell(row=k+2,column=5).value, ws_ksv.cell(row=k+2,column=6).value, ws_ksv.cell(row=k+2,column=7).value
                if ws_ksv.cell(row=k+2,column=4).value == 'Protected':
                    ws_ksv.cell(row=k+2,column=4).fill = fill_cell_green
                    counter += 1
            elif ws_ksv.cell(row=k+2,column=1).value == vm and counter != 0:
                ws_ksv.cell(row=k+2,column=1).fill = fill_cell_red
                if ws_ksv.cell(row=k+2,column=4).value == 'Protected' and ws_ksv.cell(row=k+2,column=4).value == a_1\
                   and ws_ksv.cell(row=k+2,column=5).value == b_1 and ws_ksv.cell(row=k+2,column=6).value == c_1 and ws_ksv.cell(row=k+2,column=7).value == d_1:
                    ws_ksv.cell(row=k+2,column=4).fill = fill_cell_yellow
                    ws_ksv.cell(row=k+2,column=4).value = 'NotProtected'
        vm_orig_name_dup_exception.append(vm)
                
vm_orig_name_dup_dns = []
vm_orig_name_dup_dns_exception = []
for i in range(2000):
    if ws_ksv.cell(row=i+2,column=2).value != None:
        vm_orig_name_dup_dns.append(ws_ksv.cell(row=i+2,column=2).value)
#print(len(vm_orig_name_dup_dns))

for vm in vm_orig_name_dup_dns:
    if vm_orig_name_dup_dns.count(vm) > 1 and vm not in vm_orig_name_dup_dns_exception and vm!='bxbus11':
        print(vm)
        counter = 0
        for k in range(len(vm_orig_name_dup_dns)):
            if ws_ksv.cell(row=k+2,column=2).value == vm and counter == 0:
                ws_ksv.cell(row=k+2,column=2).fill = fill_cell_red_dns
                a_1,b_1,c_1,d_1 = ws_ksv.cell(row=k+2,column=4).value, ws_ksv.cell(row=k+2,column=5).value, ws_ksv.cell(row=k+2,column=6).value, ws_ksv.cell(row=k+2,column=7).value
                if ws_ksv.cell(row=k+2,column=4).value == 'Protected':
                    ws_ksv.cell(row=k+2,column=3).fill = fill_cell_green_dns
                    counter += 1
            elif ws_ksv.cell(row=k+2,column=2).value == vm and counter != 0:
                ws_ksv.cell(row=k+2,column=2).fill = fill_cell_red_dns
                if ws_ksv.cell(row=k+2,column=4).value == 'Protected' and ws_ksv.cell(row=k+2,column=4).value == a_1\
                   and ws_ksv.cell(row=k+2,column=5).value == b_1 and ws_ksv.cell(row=k+2,column=6).value == c_1 and ws_ksv.cell(row=k+2,column=7).value == d_1:
                    ws_ksv.cell(row=k+2,column=3).fill = fill_cell_yellow_dns
                    ws_ksv.cell(row=k+2,column=4).value = 'NotProtected'
        vm_orig_name_dup_dns_exception.append(vm)
        

   
wb_ksv.save('U:\\отчеты\\audit_vm_test\\audit_vmware.xlsx')



wb_ksv = load_workbook('U:\\отчеты\\audit_vm_test\\audit_vmware.xlsx')
ws_ksv = wb_ksv['Main']
for i in range(2000):
    if ws_ksv.cell(row=i+1,column=4).value == 'Protected' and (ws_ksv.cell(row=i+1,column=5).value == None or ws_ksv.cell(row=i+1,column=6).value =='N/A'):
        ws_ksv.cell(row=i+1,column=6).value = 'Kaspersky Security для виртуальных сред. Защита без агента'
    if ws_ksv.cell(row=i+1,column=4).value == None and ws_ksv.cell(row=i+1,column=1).value != None and (ws_ksv.cell(row=i+1,column=5).value == None or ws_ksv.cell(row=i+1,column=6).value =='N/A'):
        ws_ksv.cell(row=i+1,column=4).value = 'NotProtected'
        
wb_ksv.save('U:\\отчеты\\audit_vm_test\\audit_vmware.xlsx')

wb_ksv = load_workbook('U:\\отчеты\\audit_vm_test\\audit_vmware.xlsx')
ws_ksv = wb_ksv['Main']

ws_ksv.column_dimensions['A'].width = '23'
ws_ksv.column_dimensions['B'].width = '23'
ws_ksv.column_dimensions['C'].width = '9'
ws_ksv.column_dimensions['D'].width = '12'
ws_ksv.column_dimensions['E'].width = '10'
ws_ksv.column_dimensions['F'].width = '42'
ws_ksv.column_dimensions['H'].width = '12'
ws_ksv.column_dimensions['I'].width = '13'
ws_ksv.column_dimensions['J'].width = '4'
ws_ksv.column_dimensions['K'].width = '9'
ws_ksv.column_dimensions['L'].width = '7'
ws_ksv.column_dimensions['M'].width = '28'
ws_ksv.column_dimensions['N'].width = '18'
ws_ksv.column_dimensions['O'].width = '5'
ws_ksv.column_dimensions['P'].width = '5'
ws_ksv.column_dimensions['Q'].width = '5'
ws_ksv.column_dimensions['R'].width = '5'
ws_ksv.column_dimensions['S'].width = '5'
ws_ksv.column_dimensions['T'].width = '5'


count_poweron = 0
count_poweroff = 0
count_protected = 0
count_protected_power_off = 0
count_notprotected = 0
count_protected_saas_vc_res = 0
count_poweron_saas_vc_res = 0
count_protected_saas_vc_mgmt = 0
count_poweron_saas_vc_mgmt = 0
count_poweron_saas = 0
count_poweron_saas = 0
count_poweron_iaas = 0
count_protected_iaas = 0

for i in range(2000):
    count_poweron += 1 if ws_ksv.cell(row=i+1,column=8).value=='poweredOn' else 0
    count_poweroff += 1 if ws_ksv.cell(row=i+1,column=8).value=='poweredOff' else 0
    count_protected_power_off +=1 if ws_ksv.cell(row=i+1,column=4).value =='Protected' and ws_ksv.cell(row=i+1,column=8).value=='poweredOff' else 0
    count_protected += 1 if ws_ksv.cell(row=i+1,column=4).value =='Protected' and ws_ksv.cell(row=i+1, column=4).fill != fill_cell_yellow\
                       and ws_ksv.cell(row=i+1, column=3).fill != fill_cell_yellow_dns and ws_ksv.cell(row=i+1,column=8).value=='poweredOn' else 0
    count_notprotected += 1 if ws_ksv.cell(row=i+1,column=4).value =='NotProtected' and ws_ksv.cell(row=i+1,column=8).value=='poweredOn' else 0
    
    count_protected_saas_vc_res += 1 if ws_ksv.cell(row=i+1,column=10).value == 'SaaS' and ws_ksv.cell(row=i+1,column=4).value =='Protected' and\
                                   ws_ksv.cell(row=i+1,column=8).value =='poweredOn' and ws_ksv.cell(row=i+1,column=14).value != 'vc-mgmt-01.sibintek.ru'\
                                   and ws_ksv.cell(row=i+1, column=3).fill != fill_cell_yellow_dns and ws_ksv.cell(row=i+1, column=4).fill != fill_cell_yellow  else 0
    count_poweron_saas_vc_res += 1 if ws_ksv.cell(row=i+1,column=10).value == 'SaaS' and ws_ksv.cell(row=i+1,column=8).value =='poweredOn' and ws_ksv.cell(row=i+1,column=14).value != 'vc-mgmt-01.sibintek.ru' else 0
    count_protected_saas_vc_mgmt += 1 if ws_ksv.cell(row=i+1,column=10).value == 'SaaS' and ws_ksv.cell(row=i+1,column=4).value =='Protected' and\
                                   ws_ksv.cell(row=i+1,column=8).value =='poweredOn' and ws_ksv.cell(row=i+1,column=14).value == 'vc-mgmt-01.sibintek.ru' else 0
    count_poweron_saas_vc_mgmt += 1 if ws_ksv.cell(row=i+1,column=10).value == 'SaaS' and ws_ksv.cell(row=i+1,column=8).value =='poweredOn' and ws_ksv.cell(row=i+1,column=14).value == 'vc-mgmt-01.sibintek.ru' else 0
    count_protected_iaas += 1 if ws_ksv.cell(row=i+1,column=10).value == 'IaaS' and ws_ksv.cell(row=i+1,column=4).value =='Protected'\
                            and ws_ksv.cell(row=i+1, column=3).fill != fill_cell_yellow_dns and ws_ksv.cell(row=i+1, column=4).fill != fill_cell_yellow else 0
    count_poweron_iaas += 1 if ws_ksv.cell(row=i+1,column=10).value == 'IaaS' and ws_ksv.cell(row=i+1,column=8).value =='poweredOn' else 0

    
print(count_poweron, count_poweroff, count_protected, count_notprotected)
ws_final = wb_ksv.create_sheet("Итог")
ws_final.cell(row=1,column=1).value = 'All.Serv (Включенные VM)'
ws_final.cell(row=1,column=2).value = count_poweron
ws_final.cell(row=2,column=1).value = 'Выключенные VM'
ws_final.cell(row=2,column=2).value = count_poweroff
ws_final.cell(row=3,column=1).value = 'AV.All (Под защитой включенные)'
ws_final.cell(row=3,column=2).value = count_protected
ws_final.cell(row=4,column=1).value = 'Под защитой выключенные'
ws_final.cell(row=4,column=2).value = count_protected_power_off
ws_final.cell(row=5,column=1).value = 'Незащищены включенные VM'
ws_final.cell(row=5,column=2).value = count_notprotected
ws_final.cell(row=6,column=1).value = 'SaaS включенные ВМ все инстансы'
ws_final.cell(row=6,column=2).value = (count_poweron_saas_vc_res + count_poweron_saas_vc_mgmt)
ws_final.cell(row=7,column=1).value = 'SaaS включенные ВМ vc-res-01 + hyper-v'
ws_final.cell(row=7,column=2).value = count_poweron_saas_vc_res
ws_final.cell(row=8,column=1).value = 'SaaS включенные ВМ vc-mgmt'
ws_final.cell(row=8,column=2).value = count_poweron_saas_vc_mgmt
ws_final.cell(row=9,column=1).value = 'AV.SaaS.all (SaaS включенные ВМ под защитой vc-res-01 + hyper-v)'
ws_final.cell(row=9,column=2).value = count_protected_saas_vc_res
ws_final.cell(row=10,column=1).value = 'AV.SaaS mgmt (SaaS включенные ВМ под защитой vc-mgmt)'
ws_final.cell(row=10,column=2).value = count_protected_saas_vc_mgmt
ws_final.cell(row=11,column=1).value = 'IaaS включенные ВМ'
ws_final.cell(row=11,column=2).value = count_poweron_iaas
ws_final.cell(row=12,column=1).value = 'AV.IaaS.all (IaaS включенные ВМ под защитой_'
ws_final.cell(row=12,column=2).value = count_protected_iaas


ws_final.column_dimensions['A'].width = '68'



wb_ksv.save('U:\\отчеты\\audit_vm_test\\audit_vm.xlsx')

'''

description = []
ticket = []
theme = []
with open('U:\\scripts\\itsm_changes.csv', newline='') as File:  
    reader = csv.reader(File, delimiter=';')
    for row in reader:
        #print(row)
        #print(row[3])
        check_field = ''.join(row)
        if 'исполнителейДата созданияПлановое начало' in check_field:
            for i in range(len(row)):
                if row[i] == 'Номер':
                    Number = i
                if row[i] == 'Тема':
                    Theme = i
                if row[i] == 'Описание':
                    Description = i
            continue
        description.append(row[Description])
        theme.append(row[Theme])
        ticket.append(row[Number])



wb_ksv = load_workbook('U:\\отчеты\\audit_vm_test\\audit_vmware.xlsx')
ws_ksv = wb_ksv['Main']
for m in range(len(description)):
    for i in range(2000):
        if ws_ksv.cell(row=i+1,column=1).value is not None and (ws_ksv.cell(row=i+1,column=1).value.replace('.sibintek.ru', '').lower() in description[m] or\
           ws_ksv.cell(row=i+1,column=1).value.replace('.cosn.cdc', '').lower() in description[m]):
            ws_ksv.cell(row=i+1,column=17).value = ticket[m]
        elif ws_ksv.cell(row=i+1,column=1).value is not None and (ws_ksv.cell(row=i+1,column=1).value.replace('.cosn.cdc', '').lower() in theme[m] or \
             ws_ksv.cell(row=i+1,column=1).value.replace('.sibintek.ru', '').lower() in theme[m]):
            ws_ksv.cell(row=i+1,column=18).value = ticket[m]
            
wb_ksv.save('U:\\отчеты\\audit_vm_test\\audit_vm.xlsx')



description = []
ticket = []
theme = []
with open('U:\\scripts\\itsm_task.csv', newline='') as File:  
    reader = csv.reader(File, delimiter=';')
    for row in reader:
        #print(row)
        #print(row[3])
        check_field = ''.join(row)
        if 'стоположение;Родительский объект;Мобильный телефон инициатора род. объекта;Ком' in check_field:
            for i in range(len(row)):
                if row[i] == 'Номер':
                    Number = i
                if row[i] == 'Тема':
                    Theme = i
                if row[i] == 'Описание':
                    Description = i
            continue
        description.append(row[Description])
        theme.append(row[Theme])
        ticket.append(row[Number])



wb_ksv = load_workbook('U:\\отчеты\\audit_vm_test\\audit_vm.xlsx')
ws_ksv = wb_ksv['Main']
for m in range(len(description)):
    for i in range(2000):
        if ws_ksv.cell(row=i+1,column=1).value is not None and (ws_ksv.cell(row=i+1,column=1).value.replace('.sibintek.ru', '').lower() in description[m] or\
           ws_ksv.cell(row=i+1,column=1).value.replace('.cosn.cdc', '').lower() in description[m]):
            ws_ksv.cell(row=i+1,column=19).value = ticket[m]
        elif ws_ksv.cell(row=i+1,column=1).value is not None and (ws_ksv.cell(row=i+1,column=1).value.replace('.cosn.cdc', '').lower() in theme[m] or \
             ws_ksv.cell(row=i+1,column=1).value.replace('.sibintek.ru', '').lower() in theme[m]):
            ws_ksv.cell(row=i+1,column=20).value = ticket[m]
            
wb_ksv.save('U:\\отчеты\\audit_vm_test\\audit_vm.xlsx')


'''
# Проверка с прошлым файлом
if os.path.isfile('U:\\отчеты\\audit_vm_test\\audit_vm_old.xlsx'):
    print("Checking last week audit")
    wb_ksv_old = load_workbook('U:\\отчеты\\audit_vm_test\\audit_vm_old.xlsx')

    wb_ksv_new = load_workbook('U:\\отчеты\\audit_vm_test\\audit_vm.xlsx')
    ws_ksv_new = wb_ksv_new['Main']
    ws_ksv_old = wb_ksv_old['Main']

    ws_ksv_new_vm = wb_ksv_new.create_sheet('New VM')
    ws_ksv_del_vm = wb_ksv_new.create_sheet('Deleted VM')
    ws_ksv_prot_status = wb_ksv_new.create_sheet('АВЗ')
    vm_new_name =[]
    vm_new_dns = []
    vm_new_meta = []
    vm_new_virus_status = {}
    vm_new_power = {}
    vm_old_name = []
    vm_old_dns = []
    vm_old_meta = []
    vm_old_virus_status = {}
    vm_old_power = {}
    vm_ch = []
    vm_added = []
    vm_deleted = []

    for i in range(2999):
        vm_new_name.append(ws_ksv_new.cell(row=i+2,column=1).value)
        vm_new_dns.append(ws_ksv_new.cell(row=i+2,column=2).value)
        vm_new_meta.append(ws_ksv_new.cell(row=i+2,column=3).value)
        vm_new_virus_status.update({ws_ksv_new.cell(row=i+2,column=1).value:ws_ksv_new.cell(row=i+2,column=4).value})
        vm_new_power.update({ws_ksv_new.cell(row=i+2,column=1).value:ws_ksv_new.cell(row=i+2,column=8).value})
        vm_old_name.append(ws_ksv_old.cell(row=i+2,column=1).value)
        vm_old_dns.append(ws_ksv_old.cell(row=i+2,column=2).value)
        vm_old_meta.append(ws_ksv_old.cell(row=i+2,column=3).value)
        vm_old_virus_status.update({ws_ksv_old.cell(row=i+2,column=1).value:ws_ksv_old.cell(row=i+2,column=4).value})
        vm_old_power.update({ws_ksv_old.cell(row=i+2,column=1).value:ws_ksv_old.cell(row=i+2,column=8).value})

    count_new = 0
    for i in range(len(vm_new_name)):
        if vm_new_name[i] not in vm_old_name and vm_new_meta[i] not in vm_old_meta:
            for k in range(1800):
                if ws_ksv_new.cell(row=k+1, column= 1).value == vm_new_name[i]:
                    print(k)
                    row_v = k
            for m in range(1, 26):
                ws_ksv_new_vm.cell(row=count_new+1,column=m).value = ws_ksv_new.cell(row=row_v+1, column=m).value
            count_new +=1
		
        else:
            pass
    count_old = 0
    for i in range(len(vm_old_name)):
        if vm_old_name[i] not in vm_new_name and vm_old_meta[i] not in vm_new_meta:
            for k in range(1800):
                if ws_ksv_old.cell(row=k+1, column= 1).value == vm_old_name[i]:
                    row_v = k
            for m in range(1, 26):
                ws_ksv_del_vm.cell(row=count_old+1,column=m).value = ws_ksv_old.cell(row=row_v+1, column=m).value
            count_old +=1
		
        else:
            pass
    count_virus = 0
    ws_ksv_prot_status.cell(row=1,column=1).value = 'Имя ВМ'
    ws_ksv_prot_status.cell(row=1,column=2).value = 'Новый статус'
    ws_ksv_prot_status.cell(row=1,column=3).value = 'Старый статус'
    for key, value in vm_new_virus_status.items():
        if key in vm_old_virus_status.keys() and value != vm_old_virus_status[key] :
            ws_ksv_prot_status.cell(row=count_virus + 2,column=1).value = key
            ws_ksv_prot_status.cell(row=count_virus + 2,column=2).value = value
            ws_ksv_prot_status.cell(row=count_virus + 2,column=3).value = vm_old_virus_status[key]
            count_virus += 1

    wb_ksv_old.save('U:\\отчеты\\audit_vm_test\\audit_vm_old.xlsx')
    wb_ksv_new.save('U:\\отчеты\\audit_vm_test\\audit_vm.xlsx')
else:
    pass
    print('Can\'t check last week audit. No file: audit_vm_old.xlsx')



      
try:
    Disconnect(si)
except:
    pass

try:
    os.remove("U:\\отчеты\\audit_vm_test\\audit_vmware.xlsx")
except:
    pass


print('FINISHED')
