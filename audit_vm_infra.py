from pyVim import connect
from pyVim.connect import Disconnect
import ssl
from pyVmomi import vim
import csv
import getpass
import sys
import time
import re
import getpass
import time
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Color, colors, PatternFill, Alignment
from openpyxl.styles.borders import Border, Side
import os



""" Подключение ПВ vSphere infra"""
ssl_context = ssl.SSLContext(ssl.PROTOCOL_SSLv23)
ssl_context.verify_mode = ssl.CERT_NONE
hosts = ['cdc-vc-01.infra.clouddc.ru']

# Ввод логина и пароля для дальнейшего подключения

username = input('Username (login@infra.clouddc.ru):')
password = getpass.getpass()
#print(password)
assert len(password) > 0
try :
    if len(password) < 5 or password == username:
        raise Exception
    else:
        pass
except Exception as e:
    print("Пароль не удалось скопировать")
    time.sleep(4)

#Дополнительные настройки для работы с файлами аудита

previous_audit = input("Введите имя файла аудита с расширением: ")
file_prev_audit = 'G:\\Департамент информационной безопасности\\14_ОЭСЗИ\\1_Отчеты\\АУДИТ ВМ\\script_infra\\input\\' + previous_audit
new_sheet = input("Введите текущую дату (формата хх.хх.хх)")
previous_sheet = input("Введите имя книги с прошлым аудитом: ")

# Создание функции для выгрузки данных из VMware
    
vm_metaname =[]
vm_name = []
vm_dnsname =[]
vm_power_state = []
vm_ip =[]
vm_service_type =[]
vm_type = []
vm_annotation = []
def get_vsphere_data(hosts, password, username):
    vm_metaname =[]
    vm_name = []
    vm_dnsname =[]
    vm_power_state = []
    vm_ip =[]
    vm_service_type =[]
    vm_type = []
    vm_annotation = []
    vm_os_name =[]
    for host_name in hosts:
        print('Getting data from: ' + str(host_name))
        # Подключение к кластеру
        try:
            si = connect.SmartConnect(host=host_name, user=username, pwd=password, port=443, sslContext=ssl_context)
            content = si.RetrieveContent()
            container = content.rootFolder
            viewType = [vim.VirtualMachine]
            recursive = True
            containerView = content.viewManager.CreateContainerView(container, viewType, recursive)
            children = containerView.view
        except Exception as e:
            print(e)
            pass
    # Цикл для выгрузки ВМ и работе с ними
        for vm in children:
            vm_name.append(vm.name)
            vm_power_state.append(vm.summary.runtime.powerState)
            vm_dnsname.append(vm.summary.guest.hostName)
            result_meta = (str(vm.summary.vm).replace("vim.VirtualMachine:", '')).replace("'", '')
            vm_metaname.append(result_meta)
            vm_os_name.append(vm.guest.guestFullName)
            vm_ip.append(vm.summary.guest.ipAddress)
            vm_type.append(vm.parent.name)
            vm_annotation.append(vm.config.annotation)
    vm_env= ['VMware']* len(vm_name)
    print('Amount of VM in VMware: ' + str(len(vm_name)))
    return vm_name, vm_dnsname, vm_metaname, vm_os_name, vm_ip, vm_type, vm_annotation, vm_power_state, vm_env

""" Функция для сохранения массивов в эксель"""

def save_data_to_excel(*args, filename='G:\\Департамент информационной безопасности\\14_ОЭСЗИ\\1_Отчеты\\АУДИТ ВМ\\script_infra\\input\\test.xlsx', sheet_name='Main', **kwargs):
    fill_cell_orange = PatternFill(start_color='f07233',
               end_color='f07233',
               fill_type='solid')
    alignment_cell = Alignment(horizontal='left', vertical ='center', wrapText='True')

    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
    try:
        wb_write = load_workbook(filename)
    except FileNotFoundError:
        wb_write = Workbook()
    ws_write = wb_write.create_sheet(sheet_name)
    for counter, k in enumerate(kwargs):
        #print(counter)
        #print(k)
        #print(type(k))
        ws_write.cell(row=1, column=counter+1).value = kwargs[k]
    for count_data, data in enumerate(args):
        for counter, info in enumerate(data):
            ws_write.cell(row=counter+2, column=count_data+1).value = data[counter]
            ws_write.cell(row=counter+2, column=count_data+1).alignment = alignment_cell
            ws_write.cell(row=counter+2, column=count_data+1).border = thin_border
            try:
                if 'poweredOff' in (data[counter]):
                    ws_write.cell(row=counter+2, column=count_data+1).fill = fill_cell_orange
            except:
                pass
    wb_write.save(filename)

#p = subprocess.Popen(["powershell.exe", \
#             "C:\\audit_vm\\get_hyperv_vm"],)
#p.communicate()

time.sleep(3)

'''
def get_data_from_vmm(filename='G:\\Департамент информационной безопасности\\14_ОЭСЗИ\\1_Отчеты\\АУДИТ ВМ\\script_infra\\input\\vmm_report.csv'):
    print("Reading data from Hyper-V report")
    vmm_name = []
    vmm_cr_date = []
    vmm_os_name =[]
    vmm_ip = []
    vmm_dnsname = []
    vmm_enviroment = []
    vmm_power_state = []
    vmm_type = []
    vmm_description = []
    counter = 0
    with open(filename, newline='', encoding='utf-8') as File:
        reader = csv.reader(File, delimiter =',')
        next(File)
        for row in reader:            
            if 'VirtualMachineState' in row:
                for i in range(len(row)):
                    if row[i] =='Name':
                        name = i
                    if row[i] == 'ComputerName':
                        dnsname = i
                    if row[i] == 'VirtualMachineState':
                        power = i
                    if row[i] == 'OperatingSystem':
                        os = i
                    if row[i] =='VirtualizationPlatform':
                        hyperv = i
                    if row[i] == 'ipv4Addresses':
                        ip = i
                    if row[i] =='CreationTime':
                        cr_date = i
                    if row[i] == 'Tag':
                        type_vm = i
                    if row[i] == 'Description':
                        descr = i
                continue
            #print(row[name])
            vmm_name.append(row[name])
            vmm_dnsname.append(row[dnsname])
            vmm_power_state.append(row[power])
            vmm_cr_date.append(row[cr_date])
            vmm_os_name.append(row[os])
            vmm_ip.append(row[ip])
            vmm_enviroment.append(row[hyperv])
            vmm_type.append(row[type_vm])
            vmm_description.append(row[descr])
            vmm_metaname = [None] * len(vmm_name)
    print('Amount of VM in Hyper-V: ' + str(len(vmm_name)))
    return vmm_name, vmm_dnsname, vmm_power_state, vmm_os_name, vmm_ip, vmm_cr_date, vmm_enviroment, vmm_type, vmm_description, vmm_metaname

vmm_name, vmm_dnsname, vmm_power_state, vmm_os_name, vmm_ip, vmm_cr_date, vmm_enviroment, \
                               vmm_type, vmm_description, vmm_metaname = get_data_from_vmm(filename='vmm_report.csv')
'''
# Запуск функции по выгрузке данных из VMware
vm_name, vm_dnsname, vm_metaname, vm_os_name, vm_ip, vm_type, vm_annotation, vm_power_state, vm_env = get_vsphere_data(hosts, password, username)

'''
save_data_to_excel(vm_name, vm_dnsname, vm_power_state, vm_metaname, vm_os_name, vm_ip, vm_type, vm_annotation, vm_env, vm_n='Имя ВМ', \
                   vm_dns = 'DNS имя ВМ', vm_power_state = 'Статус ВМ', vm_meta = 'Метаимя ВМ', vm_os = 'ОС', vm_i = 'IP адрес', vm_ty = 'ВМ группа', \
                   vm_ann = 'Описание', vm_e = "Платформа")

wb_ksv = load_workbook('U:\\отчеты\\audit_vm_test\\audit_infra.xlsx')
ws_ksv = wb_ksv['Main']
for i in range(len(vmm_name)):
    for m in range(800):
        if ws_ksv.cell(row=m+1, column=1).value is None:
            ws_ksv.cell(row=m+1, column=1).value = vmm_name[i].lower()
            ws_ksv.cell(row=m+1, column=2).value = vmm_dnsname[i].lower() if vmm_dnsname[i] is not None else vmm_dnsname[i]
            ws_ksv.cell(row=m+1, column=3).value = vmmm_power_state[i]
            ws_ksv.cell(row=m+1, column=4).value = vmm_metaname[i]
            ws_ksv.cell(row=m+1, column=5).value = vmm_os_name[i]
            ws_ksv.cell(row=m+1, column=6).value = vmm_ip[i]
            ws_ksv.cell(row=m+1, column=7).value = vmm_type[i]
            ws_ksv.cell(row=m+1, column=8).value = vmm_description[i]
            ws_ksv.cell(row=m+1, column=9).value = vmm_enviroment[i]
        else:
            continue
        break
wb_ksv.save('U:\\отчеты\\audit_vm_test\\audit_infra.xlsx')
'''
#Функция для работы с отчетами Касперского

def check_ksv_files(vm_name, vm_dnsname):
    ksv_status_agent =[]
    ksv_status_soft_name = []
    ksv_status_agent_sec = []
    ksv_status = []
    # Загрузка отчетов из папки
    wb_ksv = load_workbook('G:\\ksc_clouddc.xlsx')
    ws_ksv = wb_ksv['Details']
    vm_agent_dict ={}
    vm_soft_name_dict = {}
    vm_sec_name_dict = {}

    # Сравнение данных, ВМ из выгрузке, сравниваются ВМ из отчета, где совпадение ставится под защиту
    for i in range(len(vm_name)):
        for m in range(2666):
            if vm_name[i] is not None and ws_ksv.cell(row=m+1,column=3).value is not None and vm_dnsname[i] is not None and\
               (vm_dnsname[i].lower().replace('.net.cdc', '').replace('.infra.clouddc.ru', '').replace('.customers.clouddc.ru', '').replace('.infra.dc1.cdc', '').replace('.mon.cdc', '').replace('.office.cdc', '').replace('.clouddc.ru', '') ==\
                ws_ksv.cell(row=m+1,column=3).value.lower().replace('.net.cdc', '').replace('.infra.clouddc.ru', '').replace('.customers.clouddc.ru', '').replace('.infra.dc1.cdc', '').replace('.mon.cdc', '').replace('.office.cdc', '').replace('.clouddc.ru', '') or \
                vm_name[i].lower().replace('.net.cdc', '').replace('.infra.clouddc.ru', '').replace('.infra.dc1.cdc', '').replace('.customers.clouddc.ru', '').replace('.mon.cdc', '').replace('.office.cdc', '').replace('.clouddc.ru', '') ==\
                ws_ksv.cell(row=m+1,column=3).value.lower().replace('.net.cdc', '').replace('.infra.clouddc.ru', '').replace('.customers.clouddc.ru', '').replace('.infra.dc1.cdc', '').replace('.mon.cdc', '').replace('.office.cdc', '').replace('.clouddc.ru', '')):
                vm_agent_dict.update({vm_name[i]:ws_ksv.cell(row=m+1,column=4).value})
                vm_soft_name_dict.update({vm_name[i]:ws_ksv.cell(row=m+1,column=5).value})
                vm_sec_name_dict.update({vm_name[i]:ws_ksv.cell(row=m+1,column=6).value})
            elif vm_dnsname[i] is None and vm_name[i] is not None and ws_ksv.cell(row=m+1,column=3).value is not None and\
                 vm_name[i].lower().replace('.clouddc.ru', '').replace('.infra.clouddc.ru', '').replace('.infra.dc1.cdc', '').replace('.customers.clouddc.ru', '').replace('.mon.cdc', '').replace('.net.cdc', '').replace('.office.cdc', '').replace('.net.cdc', '') ==\
                 ws_ksv.cell(row=m+1,column=3).value.lower().replace('.net.cdc', '').replace('.infra.clouddc.ru', '').replace('.customers.clouddc.ru', '').replace('.infra.dc1.cdc', '').replace('.mon.cdc', '').replace('.office.cdc', '').replace('.clouddc.ru', ''):
                vm_agent_dict.update({vm_name[i]:ws_ksv.cell(row=m+1,column=4).value})
                vm_soft_name_dict.update({vm_name[i]:ws_ksv.cell(row=m+1,column=5).value})
                vm_sec_name_dict.update({vm_name[i]:ws_ksv.cell(row=m+1,column=6).value})
    wb_ksv.save('G:\\ksc_clouddc.xlsx')

    
    # Сравнение данных, ВМ из выгрузке, сравниваются ВМ из отчета, где совпадение ставится под защиту, для отчета из инфры
    
    wb_ksv = load_workbook('G:\\ksc_infra.xlsx')
    ws_ksv = wb_ksv['Details']
    for i in range(len(vm_name)):
        for m in range(2666):
            if vm_name[i] is not None and ws_ksv.cell(row=m+1,column=3).value is not None and vm_dnsname[i] is not None and\
               (vm_dnsname[i].lower().replace('.net.cdc', '').replace('.infra.clouddc.ru', '').replace('.infra.dc1.cdc', '').replace('.customers.clouddc.ru', '').replace('.mon.cdc', '').replace('.office.cdc', '').replace('.clouddc.ru', '') ==\
                ws_ksv.cell(row=m+1,column=3).value.lower().replace('.net.cdc', '').replace('.infra.clouddc.ru', '').replace('.customers.clouddc.ru', '').replace('.infra.dc1.cdc', '').replace('.mon.cdc', '').replace('.office.cdc', '').replace('.clouddc.ru', '') or\
                vm_name[i].lower().replace('.net.cdc', '').replace('.infra.clouddc.ru', '').replace('.infra.dc1.cdc', '').replace('.customers.clouddc.ru', '').replace('.mon.cdc', '').replace('.office.cdc', '').replace('.clouddc.ru', '') ==\
                ws_ksv.cell(row=m+1,column=3).value.lower().replace('.net.cdc', '').replace('.infra.clouddc.ru', '').replace('.customers.clouddc.ru', '').replace('.infra.dc1.cdc', '').replace('.mon.cdc', '').replace('.office.cdc', '').replace('.clouddc.ru', '')):
                vm_agent_dict.update({vm_name[i]:ws_ksv.cell(row=m+1,column=4).value})
                vm_soft_name_dict.update({vm_name[i]:ws_ksv.cell(row=m+1,column=5).value})
                vm_sec_name_dict.update({vm_name[i]:ws_ksv.cell(row=m+1,column=6).value})
            elif vm_dnsname[i] is None and vm_name[i] is not None and ws_ksv.cell(row=m+1,column=3).value is not None and\
                 vm_name[i].lower().replace('.net.cdc', '').replace('.infra.clouddc.ru', '').replace('.infra.dc1.cdc', '').replace('.customers.clouddc.ru', '').replace('.mon.cdc', '').replace('.office.cdc', '').replace('.clouddc.ru', '') ==\
                 ws_ksv.cell(row=m+1,column=3).value.lower().replace('.net.cdc', '').replace('.infra.clouddc.ru', '').replace('.customers.clouddc.ru', '').replace('.infra.dc1.cdc', '').replace('.mon.cdc', '').replace('.office.cdc', '').replace('.clouddc.ru', ''):
                vm_agent_dict.update({vm_name[i]:ws_ksv.cell(row=m+1,column=4).value})
                vm_soft_name_dict.update({vm_name[i]:ws_ksv.cell(row=m+1,column=5).value})
                vm_sec_name_dict.update({vm_name[i]:ws_ksv.cell(row=m+1,column=6).value})
                
    for v_2 in vm_name:
        #print(ksv_dict.get(vm_meta_name))
        ksv_status_agent.append(vm_agent_dict.get(v_2))
        ksv_status_soft_name.append(vm_soft_name_dict.get(v_2))
        ksv_status_agent_sec.append(vm_sec_name_dict.get(v_2))
        if vm_agent_dict.get(v_2) is not None and vm_sec_name_dict.get(v_2) is not None and vm_sec_name_dict.get(v_2) != 'N/A':
            ksv_status.append('Да')
        else:
            ksv_status.append('Нет')
    wb_ksv.save('G:\\ksc_infra.xlsx')
    return ksv_status, ksv_status_agent, ksv_status_soft_name, ksv_status_agent_sec


vm_name_all = vm_name
vm_dnsname_all = vm_dnsname
vm_power_state_all = vm_power_state
vm_metaname_all = vm_metaname
vm_os_name_all = vm_os_name
vm_ip_all = vm_ip
vm_type_all = vm_type
vm_annotation_all = vm_annotation
vm_env_all = vm_env

# Вызов функции сравненя с отчетами касперского

ksv_status, ksv_status_agent, ksv_status_soft_name, ksv_status_agent_sec = check_ksv_files(vm_name_all, vm_dnsname_all)
print(len(ksv_status_agent), len(vm_name_all), len(ksv_status))








'''
previous_audit = input("Введите имя файла аудита с расширением: ")
file_prev_audit = 'U:\\scripts\\infra_audit_vm\\input\\' + previous_audit
new_sheet = input("Введите текущую дату (формата хх.хх.хх)")
previous_sheet = input("Введите имя книги с прошлым аудитом: ")
'''
# Сохранение данных в excel

save_data_to_excel(vm_name_all, vm_dnsname_all, vm_power_state_all, ksv_status, ksv_status_agent, ksv_status_soft_name, ksv_status_agent_sec, vm_metaname_all,\
                   vm_os_name_all, vm_ip_all, vm_type_all, vm_annotation_all, vm_env_all, filename=file_prev_audit, sheet_name=new_sheet, vm_n='Имя ВМ', \
                   vm_dns = 'DNS имя ВМ', vm_power_state = 'Статус ВМ', ksv_status_1 = 'АВЗ', ksv_status_agent_1 = 'Версия Агента администрирования',\
                   ksv_status_soft_name_1 = 'Название программы безопасности', ksv_status_agent_sec = 'Версия программы безопасности',\
                   vm_meta = 'Метаимя ВМ', vm_os = 'ОС', vm_i = 'IP адрес', vm_ty = 'ВМ группа', \
                   vm_ann = 'Описание', vm_e = "Платформа")
'''
wb_ksv = load_workbook('U:\\отчеты\\audit_vm_test\\audit_infra.xlsx')
ws_ksv = wb_ksv['Main']
for i in range(len(vmm_name)):
    for m in range(800):
        if ws_ksv.cell(row=m+1, column=1).value is None:
            ws_ksv.cell(row=m+1, column=1).value = vmm_name[i].lower()
            ws_ksv.cell(row=m+1, column=2).value = vmm_dnsname[i].lower() if vmm_dnsname[i] is not None else vmm_dnsname[i]
            ws_ksv.cell(row=m+1, column=3).value = vmmm_power_state[i]
            ws_ksv.cell(row=m+1, column=4).value = vmm_metaname[i]
            ws_ksv.cell(row=m+1, column=5).value = vmm_os_name[i]
            ws_ksv.cell(row=m+1, column=6).value = vmm_ip[i]
            ws_ksv.cell(row=m+1, column=7).value = vmm_type[i]
            ws_ksv.cell(row=m+1, column=8).value = vmm_description[i]
            ws_ksv.cell(row=m+1, column=9).value = vmm_enviroment[i]
        else:
            continue
        break
wb_ksv.save('U:\\отчеты\\audit_vm_test\\audit_infra.xlsx')
'''
print('stop')

# Сравнение с прошлым аудитом

print("Checking last week audit")
wb_ksv = load_workbook(file_prev_audit)
ws_ksv_new = wb_ksv[str(new_sheet)]
ws_ksv_old = wb_ksv[str(previous_sheet)]
#ws_ksv_new_vm = wb_ksv_new.create_sheet('New VM')
#ws_ksv_del_vm = wb_ksv_new.create_sheet('Deleted VM')
#ws_ksv_prot_status = wb_ksv_new.create_sheet('АВЗ')
vm_new_name =[]
vm_new_dns = []
vm_new_meta = []
vm_new_virus_status = []
vm_new_power = []
vm_old_name = []
vm_old_dns = []
vm_old_meta = []
vm_old_virus_status = []
vm_old_power = []
vm_ch = []
vm_added = []
vm_deleted = []
vm_old_ip = []

fill_cell_red = PatternFill(start_color='f60511', end_color='f60511', fill_type='solid')
fill_cell_green_dns = PatternFill(start_color='52df5b', end_color='05f621', fill_type='solid')

# Цикл сравнения с прошлым аудитом

for i in range(1999):
    if ws_ksv_new.cell(row=i+2,column=1).value is not None or ws_ksv_new.cell(row=i+2,column=2).value is not None or ws_ksv_new.cell(row=i+2,column=3).value is not None:
        vm_new_name.append(ws_ksv_new.cell(row=i+2,column=1).value)
        vm_new_dns.append(ws_ksv_new.cell(row=i+2,column=2).value)
        vm_new_power.append(ws_ksv_new.cell(row=i+2,column=3).value)
    if ws_ksv_old.cell(row=i+2,column=1).value is not None or ws_ksv_old.cell(row=i+2,column=2).value is not None or ws_ksv_old.cell(row=i+2,column=3).value is not None:
        vm_old_name.append(ws_ksv_old.cell(row=i+2,column=1).value)
        vm_old_dns.append(ws_ksv_old.cell(row=i+2,column=2).value)
        vm_old_power.append(ws_ksv_old.cell(row=i+2,column=13).value)
        vm_old_ip.append(ws_ksv_old.cell(row=i+2,column=10).value)
    '''
    vm_old_virus_status.update({ws_ksv_old.cell(row=i+2,column=1).value:ws_ksv_old.cell(row=i+2,column=4).value})
    vm_old_power.update({ws_ksv_old.cell(row=i+2,column=1).value:ws_ksv_old.cell(row=i+2,column=8).value})

    '''
# Выявление новых ВМ и удаленных ВМ, в сравнении с прошлым аудитом
    
print(len(vm_old_name), len(vm_new_name), len(vm_old_dns), len(vm_new_dns))
count_new = 0
for i in range(len(vm_new_name)):
    if vm_new_name[i] not in vm_old_name and vm_new_dns[i] not in vm_old_dns:
        for k in range(1800):
            if ws_ksv_new.cell(row=k+1, column= 1).value == vm_new_name[i]:
                ws_ksv_new.cell(row=k+1,column=1).fill = fill_cell_green_dns
                break

            
for i in range(len(vm_old_name)):
    if vm_old_name[i] not in vm_new_name and vm_old_dns[i] not in vm_new_dns:
        for k in range(1800):
            if ws_ksv_new.cell(row=k+1, column= 1).value is  None and ws_ksv_new.cell(row=k+1, column= 3).value is  None\
               and ws_ksv_new.cell(row=k+1, column= 2).value is  None and ws_ksv_new.cell(row=k+1, column= 4).value is  None\
               and ws_ksv_new.cell(row=k+1, column= 5).value is  None:
                ws_ksv_new.cell(row=k+1, column= 1).value = vm_old_name[i]
                ws_ksv_new.cell(row=k+1, column= 2).value = vm_old_dns[i]
                ws_ksv_new.cell(row=k+1, column= 13).value = vm_old_power[i]
                ws_ksv_new.cell(row=k+1, column= 10).value = vm_old_ip[i]
                ws_ksv_new.cell(row=k+1,column=1).fill = fill_cell_red
                ws_ksv_new.cell(row=k+1,column=2).fill = fill_cell_red
                ws_ksv_new.cell(row=k+1, column= 3).value = 'Deleted?'
                break
		
        else:
            pass

# Список ВМ, которым не нужен АВЗ, в фале будут отмечены с графой не нужно.

exception_dict =['02-INF-CRL']
# Цикл для проставления исключений

for i in range(999):
    if ws_ksv_new.cell(row=i+2,column=1).value in exception_dict:
        ws_ksv_new.cell(row=i+2,column=4).value = 'Не нужно'






'''
for i in range(len(vm_old_name)):
    if vm_old_name[i] not in vm_new_name and vm_old_dns[i] not in vm_new_dns:
        for k in range(1800):
            if ws_ksv_old.cell(row=k+1, column= 1).value == vm_old_name[i]:
                ws_ksv_new.cell(row=k+1,column=1).fill = fill_cell_red
                break
		
        else:
            pass
    count_virus = 0

    ws_ksv_prot_status.cell(row=1,column=1).value = 'Имя ВМ'
    ws_ksv_prot_status.cell(row=1,column=2).value = 'Новый статус'
    ws_ksv_prot_status.cell(row=1,column=3).value = 'Старый статус'
    for key, value in vm_new_virus_status.items():
        if key in vm_old_virus_status.keys() and value != vm_old_virus_status[key] :
            ws_ksv_prot_status.cell(row=count_virus + 2,column=1).value = key
            ws_ksv_prot_status.cell(row=count_virus + 2,column=2).value = value
            ws_ksv_prot_status.cell(row=count_virus + 2,column=3).value = vm_old_virus_status[key]
            count_virus += 1
'''
#Сохранение итогового отчета

ter = new_sheet.split('.')
ret = '_'.join(ter)

wb_ksv.save('c:\\temp\test\\nd Antivirus_' + str(ret) + '.xlsx')

wb_ksv = load_workbook(file_prev_audit)
wb_ksv.remove(wb_ksv[new_sheet])
wb_ksv.save(file_prev_audit)

































