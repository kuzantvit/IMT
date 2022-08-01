from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Color, colors, PatternFill, Alignment
from openpyxl.styles.borders import Border, Side
import os
import csv

#file_with_audit = input("Введите полный путь и имя файла аудита с расширением: ")
file_with_audit = r"U:\Tenable\reports\audit_vm.xlsx"

#file_with_report = input("Введите путь к файлу с именем и расширением")
file_with_report = r"U:\Tenable\reports\vulns_test.csv"

def save_data_to_excel(*args, filename='G:\\Департамент информационной безопасности\\14_ОЭСЗИ\\1_Отчеты\\АУДИТ ВМ\\script_infra\\input\\test.xlsx', sheet_name='Main', **kwargs):
    fill_cell_orange = PatternFill(start_color='f07233',
               end_color='f07233',
               fill_type='solid')
    alignment_cell = Alignment(horizontal='left', vertical ='center', wrapText='True')

    thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))
    try:
        wb_write = load_workbook(filename)
    except FileNotFoundError:
        wb_write = Workbook()
    ws_write = wb_write.create_sheet(sheet_name)
    for counter, k in enumerate(kwargs):
        #print(counter)
        #print(k)
        #print(type(k))
        ws_write.cell(row=1, column=counter+1).value = kwargs[k]
    for count_data, data in enumerate(args):
        for counter, info in enumerate(data):
            ws_write.cell(row=counter+2, column=count_data+1).value = data[counter]
            ws_write.cell(row=counter+2, column=count_data+1).alignment = alignment_cell
            ws_write.cell(row=counter+2, column=count_data+1).border = thin_border
    wb_write.save(filename)

wb_audit = load_workbook(file_with_audit)
ws_audit = wb_audit['Main']





def get_data_from_report(filename=file_with_report):
    dict_1 = {}
    ten_plugin, ten_plugin_name, ten_severity, ten_netbios, ten_os, ten_netbios_compare, ten_ip, ten_repo, ten_dns_compare, ten_dns = [], [], [], [], [], [], [], [], [], []
    print("Reading tenable report")
    with open(filename, newline='', encoding='utf-8') as File:
        reader = csv.reader(File, delimiter = ',')
        for row in reader:
            if 'MAC Address' in row:
                for i in range(len(row)):
                    if row[i] =='Plugin':
                        plugin = i
                    if row[i] == 'Plugin Name':
                        plugin_name = i
                    if row[i] == 'Severity':
                        severity = i
                    if row[i] == 'Family':
                        os = i
                    if row[i] =='NetBIOS Name':
                        netbios = i
                    if row[i] == 'IP Address':
                        ip = i
                    if row[i] =='Repository':
                        repo = i
                    if row[i] == 'DNS Name':
                        dns_name = i
                continue
            ten_plugin.append(row[plugin])
            ten_plugin_name.append(row[plugin_name])
            ten_severity.append(row[severity])
            ten_os.append(row[os])
            result_netbios = row[netbios].lower().replace('cosn\\', '').replace('unknown\\', '').replace('sibintek\\', '').replace('clouddc\\', '')
            ten_netbios.append(row[netbios])
            ten_netbios_compare.append(result_netbios)
            ten_ip.append(row[ip])
            ten_repo.append(row[repo])
            result_dns =row[dns_name].lower().replace('sibintek.ru', '').replace('clouddc.ru', '').replace('infra.clouddc.ru', '').replace('cosn.cdc', '').replace('snegirsoft.com', '')
            ten_dns_compare.append(result_dns)
            ten_dns.append(row[dns_name])
            dict_1.update({row[netbios]:[ten_dns, ten_plugin_name, ten_severity, ten_os, ten_ip, ten_repo]})
    return ten_plugin, ten_plugin_name, ten_severity, ten_netbios, ten_os, ten_netbios_compare, ten_ip, ten_repo, ten_dns_compare, ten_dns, dict_1
service_model = []

ten_plugin, ten_plugin_name, ten_severity, ten_netbios, ten_os, ten_netbios_compare, ten_ip, ten_repo, ten_dns_compare, ten_dns, dict_1 =get_data_from_report()

dict_2 = {}
for i in range(2000):
    for m in range(len(ten_netbios_compare)):
        if ws_audit.cell(row =i+2, column = 1).value == ten_netbios_compare[m] or ws_audit.cell(row = i+2, column =2).value == ten_netbios_compare[m]:
            #print(ws_audit.cell(row =i+2, column = 10).value, ten_netbios_compare[m], ws_audit.cell(row =i+2, column = 1).value, ws_audit.cell(row = i+2, column =2).value)
            dict_2.update({ten_netbios[m]:ws_audit.cell(row =i+2, column = 10).value})


save_data_to_excel (ten_netbios, ten_plugin, ten_plugin_name, ten_severity, ten_os, ten_ip,\
                    ten_repo, ten_dns, filename = 'U:\\Tenable\\reports\\test.xlsx', c1 = 'Plugin')

wb_tenable = load_workbook('U:\\Tenable\\reports\\test.xlsx')
ws_tenable = wb_tenable['Main']

for i in range(500):
    for key, value in dict_2.items():
        if ws_tenable.cell(row =i+2, column =1).value == key:
            ws_tenable.cell(row=i+2, column=9).value = value
        elif ws_tenable.cell(row =i+2, column =8).value is not None and 'snegirsoft' in ws_tenable.cell(row =i+2, column =8).value:
            ws_tenable.cell(row=i+2, column=9).value = "SnegirSoft"
        elif ws_tenable.cell(row =i+2, column =8).value is not None and 'rbsv.pro' in ws_tenable.cell(row =i+2, column =8).value:
            ws_tenable.cell(row=i+2, column=9).value = "RBSV.PRO"
wb_tenable.save("U:\\Tenable\\reports\\report_final.xlsx")


